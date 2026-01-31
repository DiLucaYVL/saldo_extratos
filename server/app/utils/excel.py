from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelStyler:
    """Aplica estilos padronizados as planilhas geradas pelo sistema."""

    def __init__(self, header_color: str = "D9E1F2") -> None:
        self.header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        self.header_font = Font(bold=True, color="000000")
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def apply(self, ws, df: pd.DataFrame) -> None:
        for cell in ws[1]:
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.thin_border

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = self.thin_border

        for column in ws.columns:
            max_length = 0
            letter = column[0].column_letter
            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    continue
            ws.column_dimensions[letter].width = max_length + 2


def write_styled_excel(
    df: pd.DataFrame,
    path: Path,
    sheet_name: str = "Dados",
    header_color: str = "D9E1F2",
    styler: ExcelStyler | None = None,
) -> None:
    """Grava um DataFrame em um XLSX com estilizacao consistente."""
    if df is None:
        df = pd.DataFrame()
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] or "Dados"

    rows = dataframe_to_rows(df, index=False, header=True)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    style = styler or ExcelStyler(header_color=header_color)
    style.apply(ws, df)
    wb.save(path)


__all__ = ["write_styled_excel", "ExcelStyler"]
